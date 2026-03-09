import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import collections
import io
import re
import requests
from datetime import datetime
import xlsxwriter

st.set_page_config(page_title="RAPA Analyzer Pro", page_icon="🎪", layout="wide")

# --- CSS for better look ---
st.markdown("""
    <style>
    .main {
        background-color: #f8f9fa;
    }
    .stMetric {
        background-color: #ffffff;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    </style>
    """, unsafe_allow_html=True)

def normalize_school(name):
    if not name or str(name).lower() == 'nan': return 'Не указана'
    lower_name = str(name).lower().replace('«', '').replace('»', '').replace('"', '').strip()
    
    if lower_name in ['академия', 'академия воздушной акробатики', 'академия спорта и воздушной акробатики']:
        return 'Академия Воздушной Акробатики (объединенная)'
        
    if 'efiria' in lower_name or 'эфирия' in lower_name: return 'Эфирия'
    if 'new trend' in lower_name or 'new tend' in lower_name: return 'New Trend'
    if 'pro уровень' in lower_name or 'pro уровнь' in lower_name: return 'PRO уровень'
    if 'аверс' in lower_name: return 'Аверс'
    if 'эклипс' in lower_name: return 'Академия Эклипс'
    if 'астерия' in lower_name or 'asteria' in lower_name: return 'Центр воздушной акробатики и спорта "Астерия"'
    if 'астар' in lower_name: return 'Астар'
    if 'топ фит' in lower_name or 'top fit' in lower_name: return 'Top fit'
    if 'рожковой' in lower_name: return 'Школа Воздушной Акробатики Елены Рожковой'
    return name

def calculate_age(birthdate_str):
    if not birthdate_str: return None
    try:
        birth = datetime.strptime(str(birthdate_str).split('T')[0], '%Y-%m-%d')
        now = datetime.now()
        age = now.year - birth.year
        if (now.month, now.day) < (birth.month, birth.day):
            age -= 1
        return age
    except:
        return None

@st.cache_data
def process_excel(file_bytes):
    """
    Парсит загруженный excel-файл.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, max_col=4, values_only=True))
    
    people = []
    current_person: dict = {}
    
    for row in rows:
        participant, disc, comp, desc = row
        
        is_new_person = False
        p_str = str(participant).strip() if participant else ''
        if p_str and p_str.lower() != 'nan':
            if any(c.isalpha() for c in p_str) and not p_str.startswith('жен.,') and not p_str.startswith('муж.,') and 'лет ' not in p_str.lower() and not p_str.lower().startswith('дуэты'):
                is_new_person = True
    
        if is_new_person:
            if current_person and current_person.get('_participants'):
                # Determine perfType for the previous person/group
                p_participants = current_person.get('_participants', [])
                p_count = len(p_participants)
                perf_type = 'Соло' if p_count == 1 else ('Дуэт' if p_count == 2 else 'Группа')
                for p_sub in p_participants:
                    if isinstance(p_sub, dict):
                        people.append({**p_sub, 'Тип номера': perf_type})
            
            current_person = {
                '_participants': [],
                'Дисциплина': str(disc) if disc else '',
                'Категория': str(comp) if comp else '',
                'Ранг': 'Без разряда', # Excel doesn't always have it clear
                'Город': 'Не указан',
                'Школа': 'Не указана',
                'Тренер': 'Не указан'
            }
    
        if current_person:
            if participant and not is_new_person:
                if 'жен.,' in str(participant) or 'муж.,' in str(participant):
                    parts = str(participant).split(',')
                    age_str = [pt for pt in parts if 'лет' in pt]
                    dob = age_str[0].strip() if age_str else str(participant)
                    
                    # Logic for Excel is slightly different, but let's try to capture info
                    current_person['_participants'].append({
                        'ФИО Участника': p_str if p_str else 'Участник',
                        'Дата рождения': dob,
                        'Возраст': None, # hard to calc from "10 лет" string accurately
                        'Дисциплина': current_person['Дисциплина'],
                        'Категория': current_person['Категория'],
                        'Ранг': current_person['Ранг'],
                        'Город': current_person['Город'],
                        'Школа': current_person['Школа'],
                        'Тренер': current_person['Тренер']
                    })
            
            if desc:
                desc_str = str(desc).strip()
                if desc_str.startswith('Город:'):
                    current_person['Город'] = desc_str.replace('Город:', '').strip()
                elif desc_str.startswith('Школа'):
                    current_person['Школа'] = desc_str.split(':', 1)[-1].strip()
                elif desc_str.startswith('ФИО тренера:'):
                    current_person['Тренер'] = desc_str.replace('ФИО тренера:', '').strip()
                
                # Update added participants
                for p_sub in current_person['_participants']:
                    p_sub['Город'] = current_person['Город']
                    p_sub['Школа'] = current_person['Школа']
                    p_sub['Тренер'] = current_person['Тренер']

    if current_person:
        participants_list = current_person.get('_participants', [])
        p_count = len(participants_list)
        perf_type = 'Соло' if p_count == 1 else ('Дуэт' if p_count == 2 else 'Группа')
        for p_sub in participants_list:
            people.append({**p_sub, 'Тип номера': perf_type})
        
    for p in people:
        p['Школа'] = normalize_school(p.get('Школа', 'Не указана'))
        
    return pd.DataFrame(people), "Загружено из Excel", ""

@st.cache_data
def process_url(url):
    match = re.search(r'([0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12})', url)
    if not match:
        raise ValueError("Некорректная ссылка. Убедитесь, что она содержит ID соревнования.")
    
    comp_id = match.group(1)
    
    # Metadata
    comp_name = "Соревнование RAPA"
    comp_dates = ""
    try:
        meta_res = requests.get(f"https://referee.f-rapa.ru/api/competitions/{comp_id}/public")
        if meta_res.ok:
            meta = meta_res.json()
            comp_name = meta.get('name', comp_name)
            city = meta.get('city', {}).get('name', '')
            if city: comp_name += f" — {city}"
            
            d1 = meta.get('date_from')
            d2 = meta.get('date_to')
            if d1 and d2:
                comp_dates = f"{datetime.strptime(d1.split('T')[0], '%Y-%m-%d').strftime('%d.%m.%Y')} — {datetime.strptime(d2.split('T')[0], '%Y-%m-%d').strftime('%d.%m.%Y')}"
    except: pass

    people = []
    page = 1
    take = 50
    
    while True:
        api_url = f"https://referee.f-rapa.ru/api/competitions/{comp_id}/public/applications?page={page}&take={take}"
        resp = requests.get(api_url)
        resp.raise_for_status()
        
        data = resp.json()
        items = data.get('data', [])
        if not items: break
            
        for item in items:
            city_name = item.get('city', {}).get('name', 'Не указан') if item.get('city') else 'Не указан'
            school = normalize_school(item.get('trainer_place'))
            trainer = item.get('trainer_fullname') or 'Не указан'
            
            dcr = item.get('discipline_category_rank', {})
            rank = dcr.get('name', 'Без разряда')
            category = dcr.get('category', {}).get('name', 'Без категории')
            discipline = dcr.get('category', {}).get('discipline', {}).get('name', 'Без дисциплины')
            
            # Logic for KMS/MS
            refined_category = category
            if 'кмс' in category.lower() or 'мс' in category.lower():
                if 'кмс' in rank.lower(): refined_category = 'КМС'
                elif 'мс' in rank.lower(): refined_category = 'МС'
            
            participants = item.get('participants', [])
            perf_type = 'Соло' if len(participants) == 1 else ('Дуэт' if len(participants) == 2 else 'Группа')
            
            for part in participants:
                fullname = f"{part.get('lastname', '')} {part.get('firstname', '')} {part.get('middlename', '')}".strip()
                bdate = part.get('birthdate', '')
                age = calculate_age(bdate)
                
                people.append({
                    'ФИО Участника': fullname,
                    'Дата рождения': bdate.split('T')[0] if bdate else '',
                    'Возраст': age,
                    'Дисциплина': discipline,
                    'Категория': refined_category,
                    'Ранг': rank,
                    'Тип номера': perf_type,
                    'Город': city_name,
                    'Школа': school,
                    'Тренер': trainer
                })
                
        if not data.get('meta', {}).get('hasNextPage'): break
        page = int(page) + 1
        
    return pd.DataFrame(people), comp_name, comp_dates

def get_excel_download_link(df, comp_name, comp_dates):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    
    # Metadata rows
    meta_df = pd.DataFrame([[comp_name], [comp_dates], []])
    meta_df.to_excel(writer, index=False, header=False, sheet_name='Участники')
    
    # Data
    df.to_excel(writer, index=False, startrow=3, sheet_name='Участники')
    
    workbook = writer.book
    worksheet = writer.sheets['Участники']
    
    # Header format
    header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(3, col_num, value, header_format)
    
    # Autofilter
    worksheet.autofilter(3, 0, 3 + len(df), len(df.columns) - 1)
    
    # Column widths
    widths = [30, 15, 10, 20, 20, 15, 12, 35, 15, 25]
    for i, w in enumerate(widths):
        worksheet.set_column(i, i, w)
        
    writer.close()
    return output.getvalue()

# --- Main Logic ---

def main():
    st.title("🎪 RAPA Analyzer Pro")
    st.markdown(f"**Последнее обновление:** {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    
    data_source = st.radio("Источник:", ["Ссылка RAPA", "Excel файл"], horizontal=True)
    
    df_base = pd.DataFrame()
    c_name, c_dates = "Соревнование", ""

    if data_source == "Excel файл":
        uploaded_file = st.file_uploader("Загрузите .xlsx", type=["xlsx"])
        if uploaded_file:
            df_base, c_name, c_dates = process_excel(uploaded_file.getvalue())
    else:
        url_input = st.text_input("Ссылка на результаты:", placeholder="https://referee.f-rapa.ru/public/competition-applications/...")
        if url_input:
            try:
                df_base, c_name, c_dates = process_url(url_input)
            except Exception as e:
                st.error(f"Ошибка: {e}")

    if df_base is not None and not df_base.empty:
        st.header(f"📊 {c_name}")
        if c_dates: st.subheader(f"📅 {c_dates}")

        # --- Cascading Filters Logic ---
        st.sidebar.header("🎯 Фильтры")
        
        def get_filtered_options(df, current_filters, target_field):
            temp_df = df.copy()
            for field, val in current_filters.items():
                if val and field != target_field:
                    temp_df = temp_df[temp_df[field] == val]
            return sorted(temp_df[target_field].unique().tolist())

        # State for filters
        if 'filters' not in st.session_state:
            st.session_state.filters = {
                'Дисциплина': '', 'Категория': '', 'Ранг': '', 'Город': '', 'Школа': ''
            }

        # Sidebar inputs
        search_fio = st.sidebar.text_input("🔍 Поиск по ФИО", "")
        
        f_discip = st.sidebar.selectbox("Дисциплина", ["Все"] + get_filtered_options(df_base, st.session_state.filters, 'Дисциплина'))
        st.session_state.filters['Дисциплина'] = f_discip if f_discip != "Все" else ""
        
        f_cat = st.sidebar.selectbox("Категория", ["Все"] + get_filtered_options(df_base, st.session_state.filters, 'Категория'))
        st.session_state.filters['Категория'] = f_cat if f_cat != "Все" else ""
        
        f_rank = st.sidebar.selectbox("Ранг", ["Все"] + get_filtered_options(df_base, st.session_state.filters, 'Ранг'))
        st.session_state.filters['Ранг'] = f_rank if f_rank != "Все" else ""
        
        f_city = st.sidebar.selectbox("Город", ["Все"] + get_filtered_options(df_base, st.session_state.filters, 'Город'))
        st.session_state.filters['Город'] = f_city if f_city != "Все" else ""
        
        f_school = st.sidebar.selectbox("Школа", ["Все"] + get_filtered_options(df_base, st.session_state.filters, 'Школа'))
        st.session_state.filters['Школа'] = f_school if f_school != "Все" else ""

        if st.sidebar.button("Сбросить все фильтры"):
            st.session_state.filters = {k: '' for k in st.session_state.filters}
            st.rerun()

        # Apply all filters
        df = df_base.copy()
        if search_fio: df = df[df['ФИО Участника'].str.contains(search_fio, case=False, na=False)]
        for field, val in st.session_state.filters.items():
            if val: df = df[df[field] == val]

        # Metrics
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Заявок", len(df))
        m2.metric("Участников", df['ФИО Участника'].nunique())
        m3.metric("Школ", df['Школа'].nunique())
        m4.metric("Городов", df['Город'].nunique())

        tab1, tab2, tab3 = st.tabs(["📋 Список участников", "🏫 Статистика по школам", "🎭 Типы номеров"])

        with tab1:
            st.dataframe(df, use_container_width=True, height=500)
            excel_data = get_excel_download_link(df, c_name, c_dates)
            st.download_button(
                label="📥 Скачать Excel (с автофильтрами)",
                data=excel_data,
                file_name=f"RAPA_Analytics_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with tab2:
            st.subheader("Статистика уникальных участников по школам")
            school_stats = df.groupby(['Школа', 'Город'])['ФИО Участника'].nunique().reset_index()
            school_stats.columns = ['Школа', 'Город', 'Кол-во участников']
            st.dataframe(school_stats.sort_values('Кол-во участников', ascending=False), use_container_width=True)

        with tab3:
            st.subheader("Разделение по типам номеров")
            # To avoid double counting participants in a single duet/group, we should ideally have the App ID.
            # But we can approximate by grouping by School, City, Discipline, Category, and Rank.
            # In process_url we have rank/category/discipline.
            perf_stats = df.copy()
            # Group by 'Тип номера' and Школа
            perf_summary = perf_stats.groupby(['Школа', 'Город', 'Тип номера']).size().reset_index(name='Кол-во записей')
            # Since 'Кол-во записей' for Duo/Group counts participants, we need to divide if we want "Numbers count"
            # but for simplicity let's just show counts as is or mention it.
            st.info("💡 Примечание: Для Дуэтов и Групп количество указано по числу участников.")
            st.dataframe(perf_summary.pivot_table(index=['Школа', 'Город'], columns='Тип номера', values='Кол-во записей', fill_value=0), use_container_width=True)

    else:
        st.info("Ожидание данных для начала анализа...")


if __name__ == "__main__":
    main()
