import pandas as pd
from openpyxl import load_workbook
import collections
import sys
import os

def normalize_school(name):
    lower_name = name.lower().replace('«', '').replace('»', '').replace('"', '').strip()
    
    # Specific Perm 'Академия' cluster merged by user request
    if lower_name in ['академия', 'академия воздушной акробатики', 'академия спорта и воздушной акробатики'] or lower_name == 'академия':
        return 'Академия Воздушной Акробатики (объединенная)'
        
    if 'efiria' in lower_name or 'эфирия' in lower_name: return 'Эфирия'
    if 'new trend' in lower_name or 'new tend' in lower_name: return 'New Trend'
    if 'pro уровень' in lower_name or 'pro уровнь' in lower_name: return 'PRO уровень'
    if 'аверс' in lower_name: return 'Аверс'
    if 'эклипс' in lower_name: return 'Академия Эклипс'
    if 'астерия' in lower_name or 'asteria' in lower_name: return 'Центр воздушной акробатики и спорта "Астерия"'
    if 'астар' in lower_name: return 'Астар'
    if 'топ фит' in lower_name or 'top fit' in lower_name: return 'Top fit'
    if 'рожковой' in lower_name: return 'Школа Воздушной Акробатики Елены Рожковой'
    return name

def parse_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, max_col=4, values_only=True))
    performances = []
    current_perf = {'participants': [], 'city': 'Не указан', 'school': 'Не указана'}

    for row in rows:
        participant, disc, comp, desc = row
        
        # New performance block starts when City is found
        if desc and isinstance(desc, str) and str(desc).startswith('Город:'):
            if current_perf['participants']:
                performances.append(current_perf)
            current_perf = {'participants': [], 'city': str(desc).replace('Город:', '').strip(), 'school': 'Не указана'}
            
        # Extract School
        if desc and isinstance(desc, str):
            if str(desc).startswith('Школа (секция):'):
                current_perf['school'] = str(desc).replace('Школа (секция):', '').strip()
            elif str(desc).startswith('Школа/студия:'):
                current_perf['school'] = str(desc).replace('Школа/студия:', '').strip()
                
        # Extract Participants
        if participant and isinstance(participant, str):
            p_str = str(participant).strip()
            if p_str and p_str.lower() != 'nan':
                # Skip technical rows (categories, age, 'жен.', etc)
                if any(c.isalpha() for c in p_str) and not p_str.startswith('жен.,') and not p_str.startswith('муж.,') and not 'лет ' in p_str.lower() and not p_str.lower().startswith('дуэты'):
                    current_perf['participants'].append(p_str)

    if current_perf['participants']:
        performances.append(current_perf)

    schools_data = collections.defaultdict(lambda: {'participants': 0, 'performances': 0, 'perf_types': collections.defaultdict(int)})

    total_children = 0
    total_performances = 0

    for perf in performances:
        raw_school = perf['school'] or 'Не указана'
        school = normalize_school(raw_school)
        city = perf['city'] or 'Не указан'
        num_p = len(perf['participants'])
        
        if num_p > 0:
            schools_data[(school, city)]['participants'] += num_p
            schools_data[(school, city)]['performances'] += 1
            schools_data[(school, city)]['perf_types'][num_p] += 1
            total_children += num_p
            total_performances += 1

    print('--- ИТОГОВЫЙ ОТЧЕТ ---\n')
    print(f'Общее количество участвующих детей: {total_children}')
    print(f'Общее количество номеров: {total_performances}\n')
    print('Распределение: Школа -> Город\n')

    for (school, city), data in sorted(schools_data.items(), key=lambda x: str(x[0][0])):
        print(f'- Школа: {school} (Город: {city})')
        print(f'  Номеров: {data["performances"]}, Детей: {data["participants"]}')
        details = []
        if data['perf_types'][1]: details.append(f"{data['perf_types'][1]} (соло)")
        if data['perf_types'][2]: details.append(f"{data['perf_types'][2]} (дуэт)")
        if data['perf_types'][3]: details.append(f"{data['perf_types'][3]} (трио)")
        for k, v in data['perf_types'].items():
            if k > 3: details.append(f"{v} (группа по {k} чел)")
        if details: print(f'  Типы номеров: {", ".join(details)}')
        print('')

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python parse_excel.py <путь_к_файлу.xlsx>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"Ошибка: Файл '{file_path}' не найден.")
        sys.exit(1)
        
    parse_excel(file_path)
